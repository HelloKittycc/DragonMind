from io import BytesIO

from src.modules.knowledge.parsers.base import KnowledgeFileParser, ParserError


class XlsxKnowledgeParser(KnowledgeFileParser):
    supported_extensions = {".xlsx"}

    def extract_text(self, content: bytes, filename: str) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ParserError("XLSX parser dependency is not installed") from exc

        try:
            workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        except Exception as exc:
            raise ParserError("XLSX file could not be parsed") from exc

        sheets: list[str] = []
        for sheet in workbook.worksheets:
            lines = [f"[Sheet: {sheet.title}]"]
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = ["" if value is None else str(value).strip() for value in row]
                if any(values):
                    lines.append(f"row {row_index}: " + " | ".join(values))
            if len(lines) > 1:
                sheets.append("\n".join(lines))
        workbook.close()
        return "\n\n".join(sheets)
